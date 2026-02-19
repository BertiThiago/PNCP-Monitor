import requests
import pandas as pd
import json
import unicodedata
import os
import re
import time
from datetime import datetime, timedelta
from telegram import Bot

# ================= CONFIG =================

MAX_PAGINAS = 80
DIAS_BUSCA = 30
VALOR_MINIMO = 0
UF_FILTRO = []

VALOR_BONUS_GRANDE = 1000000
BONUS_GRANDE = 3

TIMEOUT_REQUEST = 30
MAX_RETRIES = 3

MODALIDADES = {
    1: "Concorrência",
    2: "Tomada de Preços",
    3: "Convite",
    6: "Pregão",
    7: "Dispensa",
    8: "Inexigibilidade",
    9: "RDC"
}

BASE_URL = "https://pncp.gov.br/pncp-consulta/v1/contratacoes/publicacao"

BOT_TOKEN = os.getenv("TELEGRAM_TOKEN")
CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")

# ================= FUNÇÕES =================

def normalizar(texto):
    if pd.isna(texto):
        return ""
    texto = unicodedata.normalize("NFKD", str(texto))
    return texto.encode("ASCII", "ignore").decode("ASCII").lower()

def request_com_retry(url, params):
    for tentativa in range(MAX_RETRIES):
        try:
            r = requests.get(url, params=params, timeout=TIMEOUT_REQUEST)
            if r.status_code == 200:
                return r
        except Exception:
            pass
        time.sleep(2)
    return None

def carregar_historico():
    if os.path.exists("historico_ids.json"):
        with open("historico_ids.json", "r") as f:
            return set(json.load(f))
    return set()

def salvar_historico(ids):
    with open("historico_ids.json", "w") as f:
        json.dump(list(ids), f)

def carregar_palavras():
    df = pd.read_excel("palavras_chave.xlsx")
    df.columns = df.columns.str.strip().str.lower()

    if "empresa" not in df.columns or "palavra" not in df.columns:
        raise Exception("Colunas 'empresa' e 'palavra' obrigatórias.")

    df["palavra_norm"] = df["palavra"].apply(normalizar)
    df = df[df["palavra_norm"].str.strip() != ""]

    mapa = (
        df.groupby("empresa")["palavra_norm"]
        .apply(list)
        .to_dict()
    )

    print(f"🏢 Empresas carregadas: {len(mapa)}")
    return mapa

def enviar_telegram(arquivo, mensagem):
    bot = Bot(token=BOT_TOKEN)
    bot.send_message(chat_id=CHAT_ID, text=mensagem, parse_mode="HTML")
    with open(arquivo, "rb") as f:
        bot.send_document(chat_id=CHAT_ID, document=f)

def limpar_excel(valor):
    if isinstance(valor, str):
        return re.sub(r"[\x00-\x1F\x7F]", "", valor)
    return valor

def match_avancado(descricao, palavras):
    score = 0
    for p in palavras:
        termos = p.split()
        if all(t in descricao for t in termos):
            score += 2
        elif any(t in descricao for t in termos):
            score += 1
    return score

def classificar_score(score):
    if score >= 8:
        return "🔥 ALTÍSSIMA"
    elif score >= 5:
        return "🚀 ALTA"
    elif score >= 3:
        return "⚡ MÉDIA"
    else:
        return "🟢 BAIXA"

def formatar_data(data_str):
    if not data_str:
        return ""
    try:
        return datetime.fromisoformat(data_str.replace("Z","")).strftime("%d/%m/%Y %H:%M")
    except:
        return data_str

def calcular_dias_restantes(data_enc):
    if not data_enc:
        return None
    try:
        data_obj = datetime.fromisoformat(data_enc.replace("Z",""))
        return (data_obj - datetime.now()).days
    except:
        return None

def classificar_urgencia(dias):
    if dias is None:
        return ""
    if dias < 0:
        return "⚫ ENCERRADA"
    elif dias <= 5:
        return "🔴 URGENTE"
    elif dias <= 10:
        return "🟠 ATENÇÃO"
    else:
        return "🟢 NO PRAZO"

# ================= EXECUÇÃO =================

mapa_palavras = carregar_palavras()
ids_vistos = carregar_historico()
novos_ids = set(ids_vistos)

data_final = datetime.now()
data_inicial = data_final - timedelta(days=DIAS_BUSCA)

resultados = []

for codigo_modalidade, nome_modalidade in MODALIDADES.items():

    print(f"\n🔎 Processando modalidade: {nome_modalidade}")
    pagina = 1

    while pagina <= MAX_PAGINAS:

        params = {
            "dataInicial": data_inicial.strftime("%Y%m%dT00:00:00"),
            "dataFinal": data_final.strftime("%Y%m%dT23:59:59"),
            "codigoModalidadeContratacao": codigo_modalidade,
            "pagina": pagina,
            "tamanhoPagina": 50
        }

        r = request_com_retry(BASE_URL, params)
        if not r:
            break

        dados = r.json()
        lista = dados.get("data", [])
        total_paginas = dados.get("totalPaginas", 1)

        if not lista:
            break

        for item in lista:

            descricao_original = item.get("objetoCompra", "")
            descricao = normalizar(descricao_original)
            valor = item.get("valorTotalEstimado") or 0
            uf = item.get("unidadeOrgao", {}).get("ufSigla", "")
            numero = str(item.get("numeroControlePNCP"))

            data_publicacao_raw = item.get("dataPublicacaoPncp", "")
            data_encerramento_raw = item.get("dataEncerramentoProposta", "")

            dias_restantes = calcular_dias_restantes(data_encerramento_raw)

            if UF_FILTRO and uf not in UF_FILTRO:
                continue

            if valor < VALOR_MINIMO:
                continue

            for empresa, palavras in mapa_palavras.items():

                score = match_avancado(descricao, palavras)

                if valor > VALOR_BONUS_GRANDE:
                    score += BONUS_GRANDE

                if score == 0:
                    continue

                status = "🆕 NOVA" if numero not in ids_vistos else "✔ JÁ ANALISADA"
                novos_ids.add(numero)

                resultados.append({
                    "empresa": empresa,
                    "modalidade": nome_modalidade,
                    "numero": numero,
                    "data_publicacao": formatar_data(data_publicacao_raw),
                    "data_encerramento": formatar_data(data_encerramento_raw),
                    "dias_restantes": dias_restantes,
                    "urgencia_prazo": classificar_urgencia(dias_restantes),
                    "orgao": item.get("orgaoEntidade", {}).get("razaoSocial", ""),
                    "uf": uf,
                    "objeto": descricao_original,
                    "valor": valor,
                    "score": score,
                    "prioridade_score": classificar_score(score),
                    "status": status,
                    "link_pncp": f"https://pncp.gov.br/app/editais/{numero}",
                    "link_orgao": item.get("linkSistemaOrigem","")

                })

        if pagina >= total_paginas:
            break

        pagina += 1

# ================= EXPORTAÇÃO =================

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

if resultados_por_empresa:

    nome_arquivo = f"relatorio_pncp_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    with pd.ExcelWriter(nome_arquivo, engine="openpyxl") as writer:

        resumo = []

        for empresa, dados in resultados_por_empresa.items():

            df = pd.DataFrame(dados)

            if df.empty:
                continue

            # Ordenação estratégica
            if "score" in df.columns:
                df.sort_values(by="score", ascending=False, inplace=True)

            sheet_name = empresa[:30]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]

            # ====== ESTILO CABEÇALHO ======
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # ====== AJUSTAR LARGURA AUTOMÁTICA ======
            for i, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                )
                ws.column_dimensions[get_column_letter(i)].width = min(max_length + 4, 60)

            # ====== CONGELAR PRIMEIRA LINHA ======
            ws.freeze_panes = "A2"

            # ====== HYPERLINK REAL (SEM CORROMPER EXCEL) ======
            if "link_pncp" in df.columns:
                col_link_pncp = df.columns.get_loc("link_pncp") + 1

                for row in range(2, len(df) + 2):
                    cell = ws.cell(row=row, column=col_link_pncp)
                    url = df.iloc[row-2]["link_pncp"]
                    if url:
                        cell.hyperlink = url
                        cell.style = "Hyperlink"

            if "link_orgao" in df.columns:
                col_link_orgao = df.columns.get_loc("link_orgao") + 1

                for row in range(2, len(df) + 2):
                    cell = ws.cell(row=row, column=col_link_orgao)
                    url = df.iloc[row-2]["link_orgao"]
                    if url:
                        cell.hyperlink = url
                        cell.style = "Hyperlink"

            # ====== COLORIR PRIORIDADE ======
            if "prioridade" in df.columns:
                col_prioridade = df.columns.get_loc("prioridade") + 1

                for row in range(2, len(df) + 2):
                    valor = str(ws.cell(row=row, column=col_prioridade).value)

                    if "ALTA" in valor:
                        ws.cell(row=row, column=col_prioridade).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif "MEDIA" in valor:
                        ws.cell(row=row, column=col_prioridade).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif "BAIXA" in valor:
                        ws.cell(row=row, column=col_prioridade).fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

            resumo.append({
                "Empresa": empresa,
                "Total": len(df),
                "Alta Prioridade": len(df[df["prioridade"].str.contains("ALTA", na=False)]),
                "Media Prioridade": len(df[df["prioridade"].str.contains("MEDIA", na=False)]),
                "Baixa Prioridade": len(df[df["prioridade"].str.contains("BAIXA", na=False)])
            })

        # ====== ABA RESUMO EXECUTIVO ======
        df_resumo = pd.DataFrame(resumo)
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)

        ws_resumo = writer.sheets["Resumo"]

        for col in range(1, len(df_resumo.columns) + 1):
            cell = ws_resumo.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        ws_resumo.freeze_panes = "A2"

    salvar_historico(novos_ids)


    total_geral = len(df)
    total_urgentes = len(df[df["urgencia_prazo"]=="🔴 URGENTE"])

    mensagem = f"""
<b>📊 RADAR PNCP ENTERPRISE</b>

🔎 Total oportunidades: <b>{total_geral}</b>
🔴 Urgentes (≤5 dias): <b>{total_urgentes}</b>

📅 Período analisado: últimos {DIAS_BUSCA} dias
"""

    enviar_telegram(nome_arquivo, mensagem)

else:
    print("Nenhuma oportunidade encontrada.")
